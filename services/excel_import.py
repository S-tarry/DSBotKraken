import io
import openpyxl

from database.requests import get_payout_info 
from openpyxl.styles import Alignment, Border, Side, Font



async def excel_pay_list():
    payouts = await get_payout_info()
    
    if not payouts:
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Виплати"

    headers = ["ID виплати", "ID користувача", "Ім'я користувача", "Карта користувача", "Сума виплати", "Дата виплати"]
    ws.append(headers)

    total = 0
    for p in payouts:
        ws.append([
            p.id,
            p.user_id,
            p.user.username,
            p.user.user_card,
            p.amount,
            p.payout_data.strftime("%Y-%m-%d %H:%M")
        ])
        total += p.amount

    ws.append([])
    ws.append(["", "Разом: ", total, "грн"])

    # === Форматування ===
    # 1. Вирівнювання по центру і шрифт
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=11)

    # 2. Границі для таблиці
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # 3. Ширина колонок
    column_widths = {
        "A": 15,  # ID виплати
        "B": 15,  # ID користувача
        "C": 25,  # Ім'я користувача
        "D": 25,  # Карта користувача
        "E": 15,  # Сума
        "F": 20,  # Дата
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Зберегти в пам’яті
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    return file_bytes
